"""
Step 6: Export the current canonical state of charter_pipeline.db into the
authority XLSX + a nodegoat CSV.

Replaces 06_merge_into_xlsx.py's incremental-append design (which read the
per-volume review CSVs and appended new rows into a growing copy of the
XLSX) with a pure, read-only-against-the-DB, write-only-to-export-files
snapshot: this script never mutates charter_pipeline.db, and its XLSX output
always reflects the FULL current canonical state, not a delta since the
last run. The dedup-by-id logic the old append_rows_to_sheet() needed
(guarding against re-adding a row already appended on a prior run) is no
longer needed for that reason -- the persons_authority/Places_Authority/
Charter_Data sheets are cleared and rewritten in full every run.

grantor_id/recipient_id/persons_by_role/location_written_id/
location_hearing_id/locations_mentioned_ids are NOT stored anywhere in the
database (see schema.sql's comment on the charters table) -- they are
computed here, at export time, from charter_persons/charter_places, so they
can never drift from the underlying data the way a second stored copy could.

Usage:
    python 06_export_authority.py             # exports every volume
    python 06_export_authority.py --vol 1      # a single volume only
    python 06_export_authority.py --dry-run    # preview counts, no writes

Reads:  charter_pipeline.db (canonical persons/places, unflagged charters)
Writes: CHARTER_authority_file_updated.xlsx  (sibling of AUTHORITY_FILE;
        the original is never modified)
        output/review/nodegoat_export.csv  (combined across volumes --
        safe to combine now that ids are globally collision-free, unlike
        the old per-volume files this replaces)
"""

import argparse
import csv
import shutil
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
from config import AUTHORITY_FILE, REVIEW_DIR
import db

PERSON_SHEET = "persons_authority"
PLACE_SHEET = "Places_Authority"
CHARTER_SHEET = "Charter_Data"

PERSON_SHEET_COLUMNS = [
    "person_id", "canonical_name", "variant_names", "patronymic", "occupation",
    "title", "floruit_start", "floruit_end", "gender", "associated_places",
    "notes", "sources",
]
PLACE_SHEET_COLUMNS = [
    "place_id", "canonical_name", "variant_names", "place_type",
    "coordinates_lat", "coordinates_long", "region", "district",
    "modern_equivalent", "notes", "sources",
]
CHARTER_SHEET_COLUMNS = [
    "charter_id", "shelfmark", "date", "grantor_id", "recipient_id",
    "location_written_id", "locations_mentioned_ids", "topic",
    "scribe_clues", "seal_info", "notes",
]


def _display_id_lookup(kind: str) -> dict[int, str]:
    """{pk: display_id} for every person/place, canonical or not -- charter
    references can point at a still-provisional row (e.g. an unpromoted new
    entity that hasn't been flagged for review), so this must cover both."""
    df = db.get_persons() if kind == "person" else db.get_places()
    pk_col = "person_pk" if kind == "person" else "place_pk"
    return dict(zip(df[pk_col], df["display_id"]))


def build_person_rows() -> list[dict]:
    df = db.get_persons(status="canonical")
    from person_authority import _int_or_blank
    rows = []
    for r in df.to_dict("records"):
        rows.append({
            "person_id": r["display_id"], "canonical_name": r["canonical_name"],
            "variant_names": r["variant_names"], "patronymic": r["patronymic"],
            "occupation": r["occupation"], "title": r["title"],
            "floruit_start": _int_or_blank(r["floruit_start"]),
            "floruit_end": _int_or_blank(r["floruit_end"]),
            "gender": r["gender"], "associated_places": r["associated_places"],
            "notes": r["notes"], "sources": r["sources"],
        })
    return rows


def build_place_rows() -> list[dict]:
    df = db.get_places(status="canonical")
    from place_authority import _float_or_blank
    rows = []
    for r in df.to_dict("records"):
        rows.append({
            "place_id": r["display_id"], "canonical_name": r["canonical_name"],
            "variant_names": r["variant_names"], "place_type": r["place_type"],
            "coordinates_lat": _float_or_blank(r["coordinates_lat"]),
            "coordinates_long": _float_or_blank(r["coordinates_long"]),
            "region": r["region"], "district": r["district"],
            "modern_equivalent": r["modern_equivalent"],
            "notes": r["notes"], "sources": r["sources"],
        })
    return rows


def _first_by_role_category(charter_persons: list[dict], substrings: list[str],
                              exact: str | None = None) -> dict | None:
    for cp in sorted(charter_persons, key=lambda x: x["ordinal"]):
        role = (cp["role_category"] or "").lower()
        if exact is not None and role == exact:
            return cp
        if exact is None and any(s in role for s in substrings):
            return cp
    return None


def build_charter_rows(volumes: list[int] | None) -> tuple[list[dict], int, dict]:
    """Returns (rows, n_flagged_skipped, counts). Only charters with no
    pending-review references and no parse error are included -- mirrors
    06_merge_into_xlsx.py's original skip condition exactly."""
    charters_df = db.get_charters(has_review=False)
    if volumes:
        charters_df = charters_df[charters_df["volume"].isin(volumes)]
    all_df = db.get_charters()
    if volumes:
        all_df = all_df[all_df["volume"].isin(volumes)]
    n_flagged = len(all_df) - len(charters_df)

    person_ids = _display_id_lookup("person")
    place_ids = _display_id_lookup("place")

    rows = []
    for ch in charters_df.to_dict("records"):
        cp = db.get_charter_persons(ch["charter_pk"])
        cpl = db.get_charter_places(ch["charter_pk"])

        grantor = _first_by_role_category(cp, ["issuer"])
        recipient = _first_by_role_category(cp, [], exact="recipient")
        by_role: dict[str, list[str]] = {}
        for p in sorted(cp, key=lambda x: x["ordinal"]):
            display = f"{p['extracted_name']} [{person_ids.get(p['person_pk'], '')}]" if p["person_pk"] else p["extracted_name"]
            by_role.setdefault(p["role_category"] or "unknown", []).append(display)
        persons_by_role = "; ".join(f"{role}: {', '.join(names)}" for role, names in by_role.items())

        loc_writing = next((p for p in sorted(cpl, key=lambda x: x["ordinal"]) if p["role"] == "loc.writing"), None)
        loc_hearing = next((p for p in sorted(cpl, key=lambda x: x["ordinal"]) if p["role"] == "loc.hearing"), None)
        all_place_pks = [p["place_pk"] for p in cpl if p["place_pk"]]
        locations_mentioned = "; ".join(dict.fromkeys(place_ids.get(pk, "") for pk in all_place_pks if pk in place_ids))

        location_written_id = (place_ids.get(loc_writing["place_pk"], "") if loc_writing and loc_writing["place_pk"] else "")
        location_hearing_id = (place_ids.get(loc_hearing["place_pk"], "") if loc_hearing and loc_hearing["place_pk"] else "")

        rows.append({
            "charter_id": ch["charter_id_placeholder"],
            "shelfmark": ch["shelfmark_auto"],
            "date": ch["date"],
            "grantor_id": person_ids.get(grantor["person_pk"], "") if grantor and grantor["person_pk"] else "",
            "recipient_id": person_ids.get(recipient["person_pk"], "") if recipient and recipient["person_pk"] else "",
            # single loc.writing FK column; falls back to loc.hearing if a
            # charter has no loc.writing mention (mirrors the old
            # build_charter_row()'s explicit fallback).
            "location_written_id": location_written_id or location_hearing_id,
            "locations_mentioned_ids": locations_mentioned,
            "topic": ch["subject"], "scribe_clues": ch["scribe"],
            "seal_info": ch["seal_info"], "notes": ch["notes"],
        })

    return rows, n_flagged, {"total": len(all_df), "eligible": len(charters_df)}


def build_nodegoat_rows(volumes: list[int] | None) -> list[dict]:
    charters_df = db.get_charters(has_review=False)
    if volumes:
        charters_df = charters_df[charters_df["volume"].isin(volumes)]
    person_ids = _display_id_lookup("person")
    place_ids = _display_id_lookup("place")

    rows = []
    for ch in charters_df.to_dict("records"):
        cp = db.get_charter_persons(ch["charter_pk"])
        cpl = db.get_charter_places(ch["charter_pk"])
        grantor = _first_by_role_category(cp, ["issuer"])
        recipient = _first_by_role_category(cp, [], exact="recipient")
        loc_writing = next((p for p in sorted(cpl, key=lambda x: x["ordinal"]) if p["role"] == "loc.writing"), None)
        loc_hearing = next((p for p in sorted(cpl, key=lambda x: x["ordinal"]) if p["role"] == "loc.hearing"), None)
        all_place_pks = [p["place_pk"] for p in cpl if p["place_pk"]]

        by_role: dict[str, list[str]] = {}
        for p in sorted(cp, key=lambda x: x["ordinal"]):
            by_role.setdefault(p["role_category"] or "unknown", []).append(p["extracted_name"])

        rows.append({
            "charter_id": ch["charter_id_placeholder"], "shelfmark": ch["shelfmark_auto"],
            "di_reference": ch["di_reference"], "date": ch["date"],
            "date_uncertain": ch["date_uncertain"], "doc_type": ch["doc_type"],
            "subject": ch["subject"], "outcome": ch["outcome"],
            "persons_by_role": "; ".join(f"{r}: {', '.join(n)}" for r, n in by_role.items()),
            "scribe": ch["scribe"], "scribe_source": ch["scribe_source"],
            "location_written": loc_writing["extracted_name"] if loc_writing else "",
            "location_hearing": loc_hearing["extracted_name"] if loc_hearing else "",
            "locations_mentioned_ids": "; ".join(dict.fromkeys(place_ids.get(pk, "") for pk in all_place_pks if pk in place_ids)),
            "seal_info": ch["seal_info"], "language": ch["language"],
            "grantor_id": person_ids.get(grantor["person_pk"], "") if grantor and grantor["person_pk"] else "",
            "recipient_id": person_ids.get(recipient["person_pk"], "") if recipient and recipient["person_pk"] else "",
        })
    return rows


def _replace_sheet_rows(wb, sheet_name: str, columns: list[str], rows: list[dict]) -> None:
    ws = wb[sheet_name]
    # Clear every data row (row 1 is the header), keep the sheet object/formatting.
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    for row_dict in rows:
        ws.append([row_dict.get(c, "") for c in columns])


def main():
    parser = argparse.ArgumentParser(description="Export charter_pipeline.db to the authority XLSX + nodegoat CSV.")
    parser.add_argument("--vol", type=int, action="append", help="Restrict to one volume (repeatable). Default: all.")
    parser.add_argument("--dry-run", action="store_true", help="Preview counts without writing.")
    args = parser.parse_args()
    volumes = args.vol or None

    person_rows = build_person_rows()
    place_rows = build_place_rows()
    charter_rows, n_flagged, counts = build_charter_rows(volumes)

    print(f"Canonical persons: {len(person_rows)}  |  Canonical places: {len(place_rows)}")
    print(f"Charters: {counts['eligible']} ready to export, {n_flagged} flagged (review/error -- skipped), "
          f"{counts['total']} total.")

    if args.dry_run:
        print("\n[dry-run] No files written.")
        return

    from openpyxl import load_workbook

    out_path = AUTHORITY_FILE.parent / (AUTHORITY_FILE.stem + "_updated.xlsx")
    shutil.copy2(AUTHORITY_FILE, out_path)
    print(f"\nWorking on copy: {out_path.name}")

    wb = load_workbook(out_path)
    _replace_sheet_rows(wb, PERSON_SHEET, PERSON_SHEET_COLUMNS, person_rows)
    _replace_sheet_rows(wb, PLACE_SHEET, PLACE_SHEET_COLUMNS, place_rows)
    _replace_sheet_rows(wb, CHARTER_SHEET, CHARTER_SHEET_COLUMNS, charter_rows)
    # PERSON_MERGE_PLAN is a hand-authored planning sheet nothing in the
    # pipeline ever writes to -- copied forward untouched via shutil.copy2
    # above, never regenerated here.
    wb.save(out_path)
    print(f"Wrote {len(person_rows)} person row(s), {len(place_rows)} place row(s), "
          f"{len(charter_rows)} charter row(s) to {out_path.name}")

    nodegoat_rows = build_nodegoat_rows(volumes)
    nodegoat_path = REVIEW_DIR / "nodegoat_export.csv"
    if nodegoat_rows:
        with open(nodegoat_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=list(nodegoat_rows[0].keys()))
            writer.writeheader()
            writer.writerows(nodegoat_rows)
        print(f"nodegoat export: {nodegoat_path}")


if __name__ == "__main__":
    main()
